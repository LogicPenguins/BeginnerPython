#! python3
# excel-extractor.py - This program will take input from the user and extract information from an excel spreadsheet and 
# output it into another spreadsheet of user's choice in bulk.
# DISCLAIMER: For personal convenience, this spreadsheet will be located in my xlsx_files folder inside my Excel_Mini_Projs folder
# but path is viable to change for personal use.


from openpyxl.utils import get_column_letter
import openpyxl
from pathlib import Path
import pyinputplus as pyip
import time
import logging

# Initializing logging text file
log_file = f'{Path.home()}\\small-projects-python\\Excel_Projs\\logs\\excel-extractor-logs.txt'
logging.basicConfig(filename=log_file, level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')

# Gets input from user on which excel-spreadsheet to extract from
print('What excel spreadsheet would you like to extract from? ')
file_to_extract_from = pyip.inputStr('File: ', timeout=60)
ftef_modified = f'{Path.home()}\\small-projects-python\\Excel_Projs\\xlsx_files\\{file_to_extract_from}.xlsx'

# Gets input from user on which sheet in given excel-sreadsheet to extract from
print('What sheet would you like to extract from? For default sheet, input NONE.')
sheet_from = input('Sheet: ')

if sheet_from == 'NONE':
    wb1 = openpyxl.load_workbook(ftef_modified)
    sheet1 = wb1.active
else:
    wb1 = openpyxl.load_workbook(ftef_modified)
    sheet1 = wb1[sheet_from]


# Rows and Columns must be static. Empty cells will not be skipped in the process of collecting cell data
print('How many rows would you like to extract?')
rte = pyip.inputInt('Rows: ', min=1)
print('How many columns would you like to extract?')
cte = pyip.inputInt('Columns: ', min=1)


print('These are the items that will be extracted from your designated Excel Spreadsheet: ')

# Supplied amount of rows and columns are examined and their coordinates corresponding with their values 
# are added as key value pairs in the coordinate_value_dict dictionary through each iteration
coordinate_value_dict = {}
for row in range(1, rte + 1):
    for column in range(1, cte + 1):
        coordinate = f'{get_column_letter(column)}{row}'
        coordinate_value = sheet1[coordinate].value
        coordinate_value_dict[coordinate] = coordinate_value

print(coordinate_value_dict)

# Gets input from user on which excel-spreadsheet to extract to
print('What excel spreadsheet would you like to extract to? ')
file_to_extract_to = pyip.inputStr('File: ', timeout=60)
ftet_modified = f'{Path.home()}\\small-projects-python\\Excel_Projs\\xlsx_files\\{file_to_extract_to}.xlsx'

# Gets input from user on which sheet to extract to in given excel-spreadsheet
print('What sheet would you like to extract to? For default sheet, input NONE.')
sheet_to = input('Sheet: ')

if sheet_to == 'NONE':
    wb2 = openpyxl.load_workbook(ftet_modified)
    sheet2 = wb2.active
else:
    wb2 = openpyxl.load_workbook(ftet_modified)
    sheet2 = wb2[sheet_to]

# Uses enmurate method to get index and the coordinate itself from list of coordinate_value_dict keys
# The index corresponds to the same index in the list of coordinate_value.dict.values() 
# and is assigned to the sheet's cell at [coordinate]
for index, coordinate in enumerate(coordinate_value_dict.keys()):
    sheet2[coordinate] = f'{list(coordinate_value_dict.values())[index]}'

# All changes are saved
wb2.save(ftet_modified)