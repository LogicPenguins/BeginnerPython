#! python3
# update-produce.py - Corrects costs in produce sales apreadsheet.

import openpyxl
from pathlib import Path

wb = openpyxl.load_workbook(f'{Path.home()}\\small-projects-python\\Excel_Mini_Projs\\xlsx_files\\produce_sales.xlsx')
sheet = wb['Sheet']

# The produce types and their updated prices
PRICE_UPDATES = {
    'Garlic': 3.07,
    'Celery': 1.19,
    'Lemon': 1.27
}


for row_num in range(2, sheet.max_row):
    produce_name = sheet.cell(row=row_num, column=1)
    if produce_name in PRICE_UPDATES:
        sheet.cell(row=row_num, column=2).value = PRICE_UPDATES[produce_name]


wb.save(f'{Path.home()}\\small-projects-python\\Excel_Projs\\xlsx_files\\updated_produce_sales.xlsx')