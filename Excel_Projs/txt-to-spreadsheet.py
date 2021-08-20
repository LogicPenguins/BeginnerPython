#! python3
# txt-to-spreadsheet.py - Takes text file inputs and converts lines in file to excel spreadsheet
# Note: Should change paths in code accordingly 

from openpyxl.utils import get_column_letter
import pyinputplus as pyip
from pathlib import Path
import openpyxl
import os

print('How many .txt files would you like to open?')
amount_to_open = pyip.inputInt('Amount: ', timeout=60, min=1)
print('What would you like to call your new excel spreadsheet?')
excel_name = pyip.inputStr('Name: ', timeout=60)

# Initializes workbook object to later edit active spreadsheet, assigned to the variable sheet.
wb = openpyxl.Workbook()
sheet = wb.active

# Initializes empty list to hold each text file's contents in it's own sub-list.
total_contents_lst = []

# Creates a for loop to iterate the amount of times corresponding to how many files are to be opened,
# asking for specific file name input on each iteration, checking if the file exists, and then appending 
# each line in the file to it's sublist in total_contents_lst after utilizing the readlines() method.
for file_num in range(amount_to_open):      
    text_file_to_open = pyip.inputStr('File: ')
    text_file_path = f'{Path.home()}\\small-projects-python\\Excel_Projs\\txt_files\\{text_file_to_open}'
    if os.path.exists(text_file_path):
        total_contents_lst.append([])
        text_file = open(text_file_path)
        for line in text_file.readlines():
            total_contents_lst[file_num].append(line)
    else:
        print(f'Invalid Path: {text_file_path}')

for lst_index, content_lst in enumerate(total_contents_lst):
    for content_index, content in enumerate(content_lst):
        sheet[f'{get_column_letter(lst_index + 1)}{content_index + 1}'] = content

wb.save(f'{Path.home()}\\small-projects-python\\Excel_Mini_Projs\\xlsx_files\\{excel_name}')

