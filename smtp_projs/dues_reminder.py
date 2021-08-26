#! python3
# sendDuesReminders.py - Sends emails based on their status in spreadsheet.

import os
import openpyxl
import smtplib
from openpyxl.utils import get_column_letter
from pathlib import Path

def send_reminder():
    # Open the spreadsheet and get the latest dues status.
    wb = openpyxl.load_workbook(f'{Path.home()}\\small-projects-python\\smtp_projs\\xlsx_files\\dueRecords.xlsx')
    sheet = wb['Sheet1']

    last_col = sheet.max_column
    latest_month = sheet[f'{get_column_letter(last_col)}{1}'].value

    unpaid_members = {}
    # Check each member's payment status
    for r in range(2, sheet.max_row + 1):
        payment = sheet[f'{get_column_letter(last_col)}{r}'].value
        if payment != 'paid':
            name = sheet[f'{get_column_letter(1)}{r}'].value
            email = sheet[f'{get_column_letter(2)}{r}'].value
            unpaid_members[name] = email

    # Log in to email account.
    password = os.environ.get('password')
    try:
        server = smtplib.SMTP('smtp.gmail.com', 587)
        server.ehlo()
        server.starttls()
        server.login(os.environ.get('x_email'), password)
        print('Login successful.')
    except Exception as exc:
        print(f'There was an error:\n{exc}')

    # Send out reminder emails.
    for name, email in unpaid_members.items():
        body = f'Subject: {latest_month} dues unpaid.\nDear {name},\nRecords show that you have not paid dues for {latest_month}. Please make this payment as soon as possible. Thank you!'
        print(f'Sending email to {email}...')
        sendmail_status = server.sendmail(os.environ.get('x_email'), email, body)

        if sendmail_status != {}:
            print(f'There was a problem sending email to {email}: {sendmail_status}')

    print('Tasks completed. Server shutting down.')
    server.quit()

if __name__ == '__main__':
    send_reminder()
