#! python3
# blank_row_inserter.py 
# Usage - (program name) (row to begin blanks) (how many rows to blank) (excel spreadsheet file to be executed on) 
# Note: ()'s in usage example should not be included in final output

import openpyxl
import sys
from pathlib import Path
from openpyxl.utils import get_column_letter


if len(sys.argv) == 4:
    
    start_row = int(sys.argv[1])
    blank_cells = int(sys.argv[2])
    excel_file = f'{Path.home()}\\small-projects-python\\Excel_Mini_Projs\\xlsx_files\\{sys.argv[3]}'

    wb = openpyxl.load_workbook(excel_file)
    sheet = wb.active


print(f'Modifying {excel_file}...')

for x in range(blank_cells + 1):
    for y in range(sheet.max_column):
        sheet[f'{get_column_letter(y + 1)}{x + start_row}'] = ' '


wb.save(excel_file)
print('Done.')