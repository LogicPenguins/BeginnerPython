#! python3
# spreadsheet_validator.py - Will go through every valid row and column in a spreadsheet 
# and check to see if any value is neither a collection of numbers or letters and flag that coordinate

import openpyxl
from pathlib import Path
from openpyxl.utils import get_column_letter
import logging

log_file = f'{Path.home()}\\VSCPythonProjects\\Excel_Mini_Projs\\logs\\log_file.txt'
logging.basicConfig(filename=log_file, level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')

print('Opening workbook...')
wb = openpyxl.load_workbook(f'{Path.home()}\\VSCPythonProjects\\Excel_Mini_Projs\\xlsx_files\\random.xlsx')

sheet = wb['Sheet1']

invalid_values = []
for row in range(1, sheet.max_row + 1):
    for column in range(1, sheet.max_column + 1):
        coordinate = f'{get_column_letter(column)}{row}'
        coordinate_value = sheet[coordinate].value
        logging.debug(coordinate_value)
        str_coordinate_value = str(coordinate_value)
        if str_coordinate_value.isalpha():
            pass
        elif str_coordinate_value.isnumeric():
            pass
        else:
            invalid_values.append(f'{str_coordinate_value} at Cell {coordinate} is invalid.')

for invalid_value in invalid_values:
    print(invalid_value)