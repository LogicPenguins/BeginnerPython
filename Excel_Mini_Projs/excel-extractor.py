#! python3
# excel-extractor.py - This program will take input from the user and extract information from an excel spreadsheet and 
# output it into another spreadsheet of user's choice in bulk.

from openpyxl.utils import get_column_letter
import openpyxl
from pathlib import Path
import pyinputplus as pyip
import time
import logging

# Initializing logging text file
log_file = f'{Path.home()}\\small-projects-python\\Excel_Mini_Projs\\logs\\excel-extractor-logs.txt'
logging.basicConfig(filename=log_file, level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')

# For personal convenience, this spreadsheet will be located in my xlsx_files folder inside my Excel_Mini_Projs folder
# but path is viable to change for personal use
print('What excel spreadsheet would you like to extract from? ')
file_to_extract_from = pyip.inputStr('File: ', timeout=60)
ftef_modified = f'{Path.home()}\\small-projects-python\\Excel_Mini_Projs\\xlsx_files\\{file_to_extract_from}.xlsx'

print('What sheet would you like to extract from? For default sheet, input NONE.')
sheet_from = input('Sheet: ')

if sheet_from == 'NONE':
    wb1 = openpyxl.load_workbook(ftef_modified)
    sheet1 = wb1.active
else:
    wb1 = openpyxl.load_workbook(ftef_modified)
    sheet1 = wb1[sheet_from]


print('How many rows would you like to extract?')
rte = pyip.inputInt('Rows: ', min=1)
print('How many columns would you like to extract?')
cte = pyip.inputInt('Columns: ', min=1)


print('These are the items that will be extracted from your designated Excel Spreadsheet: ')

coordinate_value_dict = {}

for row in range(1, rte + 1):
    for column in range(1, cte + 1):
        coordinate = f'{get_column_letter(column)}{row}'
        coordinate_value = sheet1[coordinate].value
        coordinate_value_dict[coordinate] = coordinate_value

print(coordinate_value_dict)

print('What excel spreadsheet would you like to extract to? ')
file_to_extract_to = pyip.inputStr('File: ', timeout=60)
ftet_modified = f'{Path.home()}\\small-projects-python\\Excel_Mini_Projs\\xlsx_files\\{file_to_extract_to}.xlsx'

print('What sheet would you like to extract to? For default sheet, input NONE.')
sheet_to = input('Sheet: ')

if sheet_to == 'NONE':
    wb2 = openpyxl.load_workbook(ftet_modified)
    sheet2 = wb2.active
else:
    wb2 = openpyxl.load_workbook(ftet_modified)
    sheet2 = wb2[sheet_to]

for index, coordinate in enumerate(coordinate_value_dict.keys()):
    sheet2[coordinate] = f'{list(coordinate_value_dict.values())[index]}'

wb2.save(ftet_modified)