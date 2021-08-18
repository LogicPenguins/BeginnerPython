#! python3
# cell_inverter.py - Inverts the rows and columns in a spreadsheet.

from openpyxl.utils import get_column_letter
import pyinputplus as pyip
from pathlib import Path
import openpyxl
import time


file1 = pyip.inputStr('File: ', timeout=60)
sheet_to_open = pyip.inputStr('Sheet: ', timeout=60)

file1_path = f'{Path.home()}\\small-projects-python\\Excel_Mini_Projs\\xlsx_files\\{file1}'

# Loads given workbook to invert from.
wb1 = openpyxl.load_workbook(file1_path)
main_sheet1 = wb1[sheet_to_open]

values_list = []

# Each outer list in values_list will be a column, each inner list containing a value in respective column.
for y in range(main_sheet1.max_column):
    values_list.append([])
    for x in range(main_sheet1.max_row):
        cell_value = main_sheet1[f'{get_column_letter(y + 1)}{x + 1}'].value
        values_list[y].append(cell_value)

print('Processing data...')
time.sleep(1)
wb1.close()

file2 = pyip.inputStr('New File Name: ')
file2_path = f'{Path.home()}\\small-projects-python\\Excel_Mini_Projs\\xlsx_files\\{file2}'

print('Opening new spreadsheet...')
time.sleep(1)

# Opens a fresh excel spreadsheet.
wb2 = openpyxl.Workbook()
main_sheet2 = wb2.active

# Loops through each column in the values_list, each column being it's own list of values under that column. In this loop, x and y are switched 
# in opposite positions so that the values are inverted in comparison to the original spreadsheet.
for column in range(len(values_list)):
    for index, value in enumerate(values_list[column]):
        main_sheet2[f'{get_column_letter(index + 1)}{column + 1}'] = value

# The spreadsheet is saved with user-given name.
wb2.save(file2_path)


