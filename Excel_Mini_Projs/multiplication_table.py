#! python3
# multiplication_table.py - Creates a multiplication on an excel spreadsheet.

import openpyxl, sys
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from pathlib import Path
import os

if len(sys.argv) == 2:
    # Assigns the number provided by user to the number used as the max in multiplication table.
	try:
		num = int(sys.argv[1])
	except Exception as e:
		print(e)

    # Workbook is opened and directed to the active sheet.
	wb = openpyxl.Workbook()
	sheet = wb.active
	
	for y in range(num + 1):
		for x in range(num + 1):
		
			# Check if in header row or column.
			is_header = False
			if x == 0 and y == 0:
				is_header = True
				value = ''
			elif x == 0:
				is_header = True
				value = y			
			elif y == 0:
				is_header = True
				value = x			
			else:
				value = x * y
			coordinate = sheet[f'{get_column_letter(y + 1)}{x + 1}']	

			# Sets font of header numbers to bold.
			if is_header:
				coordinate.font = Font(bold = True)		

			# Sets values of each coordinate accordingly.	
			coordinate.value = value
	
	# Creating path for file to be saved in. Path will vary for different systems.
	file = f'{Path.home()}\\small-projects-python\\Excel_Mini_Projs\\xlsx_files\\multiplication_table_{num}.xlsx'
	wb.save(file)	
	print(f'Saved as {os.path.basename(file)}.')
else:
	print('Please include a number in your argument.')