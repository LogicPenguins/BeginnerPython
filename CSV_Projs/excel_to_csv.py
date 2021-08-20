#! python3
# excel_to_csv.py - Converts excel-spreadsheets into CSV files.

from pathlib import Path
import os, openpyxl, csv
from openpyxl.utils import get_column_letter
import logging
import time

logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')

dir_path = f'{Path.home()}\\small-projects-python\\CSV_Projs\\sample_folder'

for file in os.listdir(dir_path):
    if file.endswith('.xlsx') != True:
        continue
    file_path = f'{Path.home()}\\small-projects-python\\CSV_Projs\\sample_folder\\{file}'
    wb = openpyxl.load_workbook(file_path)
            
    for sheet_index, sheet in enumerate(wb.sheetnames):
        values_list = []
        name = wb.sheetnames[sheet_index]
        current_sheet = wb[name]

        for row in range(current_sheet.max_row):
            values_list.append([])
            for column in range(current_sheet.max_column):
                cell_value = current_sheet[f'{get_column_letter(column + 1)}{row + 1}'].value
                values_list[row].append(cell_value)
            
        wb.close()
        csv_file_path = f'{Path.home()}\\small-projects-python\\CSV_Projs\\csv_destination\\{file}_{name}.csv'
        csv_file = open(csv_file_path, 'w', newline='')
        csv_file_writer = csv.writer(csv_file)
        for row in values_list:
            output = []
            for value in row:
                output.append(value)
            csv_file_writer.writerow(output)
        
        csv_file.close()
        print(f'Finished creating {file}_{name}.csv')